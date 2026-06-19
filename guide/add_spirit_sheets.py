import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOOLS = r'C:\Users\micha\Desktop\pocketSummonerAlpha\tools'
GUIDE = r'C:\Users\micha\Desktop\pocketSummonerAlpha\guide\SpiritSummoner_Guide.xlsx'

spirits = json.load(open(f'{TOOLS}\\dump_spirits.json', encoding='utf-8'))
encounters = json.load(open(f'{TOOLS}\\spirit_encounters.json', encoding='utf-8'))

HEADER_FILL = PatternFill('solid', start_color='1A1A24')
HEADER_FONT = Font(name='Arial', bold=True, color='F2C86B')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='F2C86B')
BODY_FONT = Font(name='Arial', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOTFOUND_FILL = PatternFill('solid', start_color='FBE3E3')
FOUND_FILL = PatternFill('solid', start_color='E3F7E3')

def fmt_locations(sid):
    locs = encounters.get(sid, [])
    if not locs:
        return ''
    parts = []
    for area, task, level, drop in locs:
        if drop and drop > 0:
            parts.append(f"{area} — \"{task}\" (Lv.{level}, {int(drop*100)}% catch)")
        else:
            parts.append(f"{area} — \"{task}\" (Lv.{level}, not catchable)")
    return '\n'.join(parts)

wb = load_workbook(GUIDE)
for name in ('Spirit Roster', 'Evolution Lines'):
    if name in wb.sheetnames:
        del wb[name]

# ---------- SPIRIT ROSTER ----------
ws = wb.create_sheet('Spirit Roster')
ws['A1'] = 'Spirit Roster — Where to Find Each Spirit'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:I1')

cols = ['ID', 'Name', 'Type 1', 'Type 2', 'Category', 'Evolves From', 'Evolves Into', 'Where to Catch / Fight', 'Notes']
for i, h in enumerate(cols, start=1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal='center')

ids_sorted = sorted(spirits.keys(), key=lambda x: int(x))
id_to_name = {sid: spirits[sid]['Name'] for sid in ids_sorted}

r = 4
for sid in ids_sorted:
    v = spirits[sid]
    pre = v['PreEvolution']
    evo = v['Evolution']
    pre_name = id_to_name.get(str(pre), '') if pre else ''
    evo_name = id_to_name.get(str(evo), '') if evo else ''
    locs = fmt_locations(sid)
    if not locs and not pre:
        continue  # not catchable wild and not obtainable via evolution
    note = ''
    if not locs:
        note = f'Not found wild — obtained by evolving {pre_name}'
    row = [
        int(sid), v['Name'], v['Type1'], v['Type2'] if v['Type2'] != 'None' else '-',
        v.get('Category') or '-', pre_name or '-', evo_name or '-', locs or '-', note
    ]
    for ci, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = WRAP
    ws.cell(row=r, column=8).fill = FOUND_FILL if locs else NOTFOUND_FILL
    r += 1

widths = [6, 18, 10, 10, 12, 16, 16, 46, 36]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A4'

# ---------- EVOLUTION LINES ----------
ws2 = wb.create_sheet('Evolution Lines')
ws2['A1'] = 'Evolution Lines — Requirements per Stage'
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:G1')

cols2 = ['From Spirit', 'Evolves Into', 'Level Required', 'Item / Orb Required', 'Gold Cost', 'Cores Required', 'Notes']
for i, h in enumerate(cols2, start=1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal='center')

r = 4
for sid in ids_sorted:
    v = spirits[sid]
    evo = v['Evolution']
    if not evo:
        continue
    evo_name = id_to_name.get(str(evo), f'#{evo}')
    req = v['Requirements']['EvolveRequirements']
    item = req['ItemRequired'] or '-'
    amt = req['Amount']
    gold_cost = amt if item == 'gold' else 0
    item_display = '-' if item in ('', 'gold', None) else item
    note = ''
    if item == 'gold' and amt:
        note = f'Pure gold cost evolution (no orb item)'
    elif not item:
        note = 'No item/orb required'
    row = [
        f"{v['Name']} (#{sid})", f"{evo_name} (#{evo})", req['LevelRequired'] or '-',
        item_display, gold_cost if gold_cost else '-', req['CoresRequired'] or '-', note
    ]
    for ci, val in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=ci, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = WRAP
    r += 1

widths2 = [22, 22, 14, 26, 12, 14, 32]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A4'

wb.save(GUIDE)
print('saved, rows:', r)
