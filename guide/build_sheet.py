from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', start_color='1A1A24')
HEADER_FONT = Font(name='Arial', bold=True, color='F2C86B')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='F2C86B')
BODY_FONT = Font(name='Arial', size=10)
FIGHT_FILL = PatternFill('solid', start_color='FFF3D6')
DUEL_FILL = PatternFill('solid', start_color='EEE3FB')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Columns: Task #, Task Name, Type, Repeatable, Steps, Energy Cost, Enemy, XP, Gold, Unlocks After
COLS = ['Task #', 'Task Name', 'Type', 'Repeatable', 'Steps', 'Energy Cost', 'Enemy', 'XP Reward', 'Gold Reward', 'Unlocks After']
WIDTHS = [8, 36, 10, 11, 7, 11, 30, 11, 12, 30]

# row format: [num, name, type, repeatable, steps, energy, enemy, xp, gold, unlocks_after]
AREAS = {
    'Basalon Town': {
        'level': 1,
        'rows': [
            [1, 'Visit the Summoner Guild', 'Explore', 'No', 1, 1, '-', 2, 10, 'None — starting task'],
            [2, 'Join the Summoner Guild', 'Explore', 'No', 3, 1, '-', 2, 12, 'Visit the Summoner Guild'],
            [3, 'Training', 'Explore', 'No', 5, 1, '-', 3, 12, 'Join the Summoner Guild'],
            [4, 'Assisting in the laboratory', 'Collect', 'No', 5, 1, '-', 5, 15, 'Training'],
            [5, 'Patrol Task', 'Explore', 'No', 4, 1, '-', 5, 13, 'Assisting in the laboratory'],
            [6, 'Helping Villagers', 'Explore', 'Yes', 10, 1, '-', 6, 30, 'Patrol Task'],
            [7, 'Listening to residents problems', 'Explore', 'Yes', 5, 1, '-', 10, 10, 'Patrol Task'],
            [8, 'Catching a rat', 'Fight', 'No', 1, 1, 'Rat spirit Lv.1', 10, 50, 'Listening to residents problems'],
            [9, 'Exploring the mystic house', 'Explore', 'Yes', 10, 1, '-', 10, 10, 'Listening to residents problems'],
            [10, 'Ghost!', 'Fight', 'No', 1, 1, 'Ghost spirit Lv.5', 25, 50, 'Exploring the mystic house'],
            [11, 'Pet the dog?', 'Fight', 'No', 1, 1, 'Dogi Lv.10', 30, 55, 'Ghost!'],
            [12, "Rumor of the Dragon's Lair", 'Hear', 'No', 1, 1, '-', 0, 0, 'Dark Shire: Defeat the mage leader'],
            [13, 'Report of dark matter collection', 'Hear', 'Yes', 1, 1, '-', 0, 0, 'Dark Shire: Defeat the mage leader'],
            [14, 'Reward for completed task', 'Receive', 'No', 1, 1, '-', 500, 3000, 'Spirit Sands: The Dark Summoner retreats'],
            [15, 'Report of dark matter (2)', 'Hear', 'Yes', 1, 1, '-', 0, 0, 'Reward for completed task'],
        ]
    },
    'Dark Forest': {
        'level': 5,
        'rows': [
            [1, 'Exploring the entrance', 'Explore', 'No', 10, 1, '-', 15, 30, 'None — starting task'],
            [2, 'Catching a spirit with magical skills', 'Fight', 'No', 1, 1, 'Trifish (B) Lv.1', 66, 200, 'Exploring the entrance'],
            [3, 'Exploring the forest', 'Explore', 'Yes', 5, 1, '-', 22, 34, 'Catching a spirit with magical skills'],
            [4, 'Trifish found!', 'Fight', 'No', 1, 1, 'Trifish (B) Lv.5', 69, 250, 'Exploring the forest'],
            [5, 'Light up the path', 'Explore', 'Yes', 10, 1, '-', 21, 45, 'Trifish found!'],
            [6, 'Exploring the deep forest', 'Explore', 'Yes', 10, 1, '-', 22, 46, 'Light up the path'],
            [7, 'Check the mystic light', 'Explore', 'No', 10, 1, '-', 20, 48, 'Exploring the deep forest'],
            [8, '[Rare] BULBHEAD found!', 'Fight', 'No', 1, 1, 'Bulbhead Lv.10', 70, 280, 'Check the mystic light'],
            [9, '[Rare] A stray Koki appears', 'Fight', 'No', 1, 1, 'Koki Lv.15', 70, 100, 'Check the mystic light'],
            [10, 'SNADLE approaches!', 'Fight', 'No', 1, 1, 'Snadle Lv.12', 70, 100, 'Check the mystic light'],
            [11, 'Check the bushes...', 'Explore', 'Yes', 3, 1, '-', 25, 15, 'SNADLE approaches!'],
            [12, 'BIGMOUTH chomps!', 'Fight', 'No', 1, 1, 'Bigmouth Lv.12', 70, 100, 'Check the bushes...'],
            [13, 'Duel test', 'Duel', 'Yes', 1, 2, 'Queen Morwyn (3 spirits Lv.40)', 2000, 100, 'None (gated by player Lv.40 instead)'],
            [14, 'Dark matter collection', 'Patrol', 'Yes', 10, 1, '-', 280, 280, 'Basalon Town: Report of dark matter collection'],
            [15, 'Expelling guard spirit', 'Fight', 'No', 1, 2, 'Dark spirit Lv.20', 500, 500, 'Dark matter collection'],
        ]
    },
    'Spirit Sands': {
        'level': 8,
        'rows': [
            [1, 'Prepare water', 'Explore', 'Yes', 10, 1, '-', 35, 12, 'None — starting task'],
            [2, 'Exploring the desert', 'Explore', 'Yes', 10, 1, '-', 35, 15, 'Prepare water'],
            [3, 'Helping a lost traveler', 'Explore', 'Yes', 10, 1, '-', 37, 16, 'Exploring the desert'],
            [4, 'Big Worm!', 'Fight', 'No', 1, 1, 'Giant Worm Lv.12', 250, 400, 'Exploring the desert'],
            [5, 'Search for treasure', 'Explore', 'Yes', 10, 1, '-', 35, 20, 'Helping a lost traveler'],
            [6, 'Entering the Pyramid', 'Explore', 'Yes', 10, 1, '-', 45, 21, 'Search for treasure'],
            [7, 'Exploring the pyramid', 'Explore', 'Yes', 10, 1, '-', 55, 22, 'None listed'],
            [8, 'Investigating the heat source', 'Explore', 'No', 3, 1, '-', 68, 100, 'None listed'],
            [9, 'TRIFISH in the desert', 'Fight', 'No', 1, 1, 'Trifish (R) Lv.20', 300, 300, 'Investigating the heat source'],
            [10, 'Dark matter collection', 'Stop', 'No', 10, 1, '-', 0, 0, 'Basalon Town: Report of dark matter collection'],
            [11, 'Expelling guard spirit', 'Fight', 'No', 1, 2, 'Dark spirit Lv.20', 500, 500, 'Dark matter collection'],
            [12, 'The dark summoner', 'Talk', 'No', 1, 1, '-', 500, 500, 'Expelling guard spirit'],
            [13, 'Defeat the Dark Summoner', 'Duel', 'No', 1, 3, 'Dark Summoner (3 spirits Lv.20)', 5000, 10000, 'Expelling guard spirit'],
            [14, 'The Dark Summoner retreats', 'Talk', 'No', 1, 1, '-', 500, 1000, 'Defeat the Dark Summoner'],
        ]
    },
    'Dark Shire': {
        'level': 12,
        'rows': [
            [1, 'Break in', 'Explore', 'Yes', 10, 1, '-', 100, 205, 'None — starting task'],
            [2, 'Encounter!', 'Fight', 'No', 1, 1, 'Dark spirit Lv.25', 300, 500, 'Break in'],
            [3, 'Capture the gate', 'Explore', 'Yes', 10, 1, '-', 130, 200, 'Encounter!'],
            [4, 'Securing the gate', 'Explore', 'Yes', 10, 1, '-', 131, 190, 'Capture the gate'],
            [5, 'Expelling dark mages', 'Fight', 'No', 1, 1, 'Dark spirit Lv.28', 500, 500, 'Securing the gate'],
            [6, 'Capture the village', 'Explore', 'Yes', 10, 1, '-', 150, 160, 'None — starting task'],
            [7, 'Defeat the mage leader', 'Fight', 'No', 1, 1, 'Dark spirit Lv.30', 550, 550, 'Capture the village'],
        ]
    },
    "King's Heart Forest": {
        'level': 20,
        'rows': [
            [1, 'Camp Preparations', 'Explore', 'Yes', 10, 1, '-', 250, 21, 'None — starting task'],
            [2, 'Collect King forest wood', 'Explore', 'Yes', 10, 1, '-', 260, 20, 'Camp Preparations'],
            [3, 'Crafting artifact', 'Explore', 'Yes', 10, 1, '-', 330, 16, 'Collect King forest wood'],
            [4, 'Activate artifact', 'Explore', 'Yes', 10, 1, '-', 300, 18, 'Crafting artifact'],
            [5, 'Plant attack', 'Fight', 'No', 1, 1, 'Plant spirit Lv.10', 300, 78, 'Activate artifact'],
            [6, 'Exploring the hidden path', 'Explore', 'Yes', 10, 1, '-', 350, 0, 'Plant attack'],
            [7, 'Group!', 'Fight', 'No', 1, 1, 'Spirit #36 Lv.15', 350, 80, 'Exploring the hidden path'],
            [8, 'Breakthrough', 'Fight', 'No', 1, 1, 'Spirit #36 Lv.25', 310, 110, 'Group!'],
            [9, 'Entering the high ground', 'Explore', 'Yes', 10, 1, '-', 300, 17, 'Breakthrough'],
            [10, 'Minions', 'Fight', 'No', 1, 1, 'Dogia Lv.25', 320, 25, 'Entering the high ground'],
            [11, 'Leader', 'Fight', 'No', 1, 1, 'Steelha Lv.30', 350, 130, 'Minions (also requires player Lv.40)'],
        ]
    },
    'The Great Land': {
        'level': 25,
        'rows': [
            [1, 'Traveling to the Great Land', 'Explore', 'Yes', 10, 1, '-', 400, 66, 'None — starting task'],
            [2, 'Helping a Traveler', 'Explore', 'Yes', 10, 1, '-', 400, 67, 'Traveling to the Great Land'],
            [3, 'Escort the traveler', 'Explore', 'Yes', 10, 1, '-', 410, 68, 'Helping a Traveler'],
            [4, 'Enter the hidden lair', 'Explore', 'Yes', 10, 1, '-', 420, 69, 'Helping a Traveler'],
            [5, 'Protect the traveler', 'Fight', 'No', 1, 1, 'Foorse (O) Lv.28', 500, 68, 'Escort the traveler'],
            [6, 'Accepting reward from traveler', 'Accept', 'No', 1, 1, '-', 400, 100, 'Protect the traveler'],
            [7, 'Clearing the path', 'Fight', 'Yes', 1, 1, 'Foorse (O) Lv.30', 500, 130, 'None listed'],
            [8, 'Opening the hidden treasure', 'Explore', 'Yes', 10, 1, '-', 440, 100, 'Clearing the path'],
            [9, 'Big Bird?', 'Fight', 'Yes', 1, 1, 'Spirit #260 Lv.30', 540, 220, 'Opening the hidden treasure'],
        ]
    },
    'Sky Tower': {
        'level': 30,
        'rows': [
            [1, 'Mineral Collection', 'Collect', 'Yes', 10, 1, '-', 680, 120, 'None — starting task'],
            [2, 'Crafting key for entrance', 'Craft', 'No', 1, 2, '-', 650, 100, 'Mineral Collection'],
            [3, 'Entering the Sky Tower', 'Enter', 'Yes', 10, 1, '-', 660, 130, 'Crafting key for entrance'],
            [4, 'Battle door guard', 'Fight', 'No', 1, 1, 'Spirit #45 Lv.40', 640, 130, 'Entering the Sky Tower'],
            [5, 'Sky Tower F/1', 'Explore', 'No', 10, 1, '-', 660, 140, 'Battle door guard'],
            [6, 'Sky Tower F/2', 'Explore', 'No', 10, 1, '-', 680, 160, 'Sky Tower F/1'],
            [7, 'Guard', 'Fight', 'No', 1, 1, 'Spirit #45 Lv.40', 700, 230, 'Sky Tower F/2'],
            [8, 'Sky Tower F/3', 'Explore', 'Yes', 10, 1, '-', 630, 260, 'Guard'],
            [9, 'Sky Tower F/4', 'Explore', 'Yes', 10, 1, '-', 650, 280, 'Sky Tower F/3'],
            [10, 'Sky Tower F/5', 'Explore', 'Yes', 10, 1, '-', 650, 280, 'Sky Tower F/4'],
            [11, 'The Four Chairs (1/4) (WIP)', 'Fight', 'Yes', 1, 1, '- (unreleased content)', 0, 0, 'Sky Tower F/5'],
        ]
    },
    "The Dragon's Lair": {
        'level': 32,
        'rows': [
            [1, 'Zone of fire', 'Explore', 'Yes', 10, 1, '-', 800, 80, 'None — starting task'],
            [2, 'Fire brood approaches', 'Fight', 'No', 1, 1, 'Spirit #243 Lv.40', 0, 100, 'Zone of fire'],
            [3, 'Zone of water', 'Explore', 'Yes', 10, 1, '-', 860, 120, 'Fire brood approaches'],
            [4, 'Water brood approaches', 'Fight', 'No', 1, 1, 'Spirit #246 Lv.40', 0, 0, 'Zone of water'],
            [5, 'Zone of wind', 'Explore', 'Yes', 10, 1, '-', 880, 130, 'Water brood approaches'],
            [6, 'Wind brood approaches', 'Fight', 'No', 1, 1, 'Spirit #249 Lv.40', 0, 0, 'Zone of wind'],
            [7, 'Zone of ground', 'Explore', 'Yes', 10, 1, '-', 900, 130, 'Wind brood approaches'],
            [8, 'Ground brood approaches', 'Fight', 'No', 1, 1, 'Worm Lv.40', 0, 0, 'Zone of ground'],
        ]
    },
    'Wall of Gaia': {
        'level': 32,
        'rows': [
            [1, 'Entering the valley of Gaia', '-', 'Yes', 10, 1, '-', 700, 250, 'None — starting task'],
            [2, 'Exploring the dino base', 'Explore', 'Yes', 10, 1, '-', 700, 260, 'None listed'],
            [3, 'Defend yourself!', 'Fight', 'No', 1, 1, 'Rexli Lv.35', 900, 220, 'Exploring the dino base'],
            [4, 'Treasure!', 'Collect', 'No', 1, 2, '-', 780, 1000, 'Defend yourself!'],
            [5, 'Revenge of the dark mage', 'Fight', 'No', 1, 1, 'Spirit #33 Lv.40', 770, 500, 'Defend yourself!'],
            [6, 'Up on the mountain', 'Explore', 'Yes', 10, 1, '-', 780, 203, 'Defend yourself!'],
            [7, 'CHIKI', 'Fight', 'No', 1, 1, 'Spirit #26 Lv.40', 0, 0, 'Up on the mountain'],
            [8, 'Into the dark village', 'Fight', 'No', 1, 1, '- (no opponent listed)', 0, 0, 'Revenge of the dark mage'],
            [9, 'Keeper of the village', 'Fight', 'Yes', 1, 1, 'Spirit #38 Lv.45', 0, 0, 'Into the dark village'],
        ]
    },
}

wb = Workbook()

ov = wb.active
ov.title = 'Overview'
ov['A1'] = 'Spirit Summoner — Story Quest Guide'
ov['A1'].font = TITLE_FONT
ov.merge_cells('A1:D1')

headers = ['Area', 'Level Required', 'Tasks Documented', 'Notes']
for i, h in enumerate(headers, start=1):
    c = ov.cell(row=3, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

r = 4
for name, data in AREAS.items():
    ov.cell(row=r, column=1, value=name).font = BODY_FONT
    ov.cell(row=r, column=2, value=data['level']).font = BODY_FONT
    ov.cell(row=r, column=3, value=len(data['rows'])).font = BODY_FONT
    ov.cell(row=r, column=4, value='').font = BODY_FONT
    for col in range(1, 5):
        ov.cell(row=r, column=col).border = BORDER
    r += 1

ov.column_dimensions['A'].width = 22
ov.column_dimensions['B'].width = 14
ov.column_dimensions['C'].width = 16
ov.column_dimensions['D'].width = 55

for name, data in AREAS.items():
    ws = wb.create_sheet(name[:31])
    ws['A1'] = f"{name} — Level {data['level']}+"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells(f'A1:{get_column_letter(len(COLS))}1')

    for i, h in enumerate(COLS, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal='center')

    r = 4
    for row in data['rows']:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if ci in (8, 9):
                cell.number_format = '#,##0'
        task_type = row[2]
        fill = FIGHT_FILL if task_type == 'Fight' else (DUEL_FILL if task_type == 'Duel' else None)
        if fill:
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=r, column=ci).fill = fill
        r += 1

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

wb.save(r'C:\Users\micha\Desktop\pocketSummonerAlpha\guide\SpiritSummoner_Guide.xlsx')
print('saved')
